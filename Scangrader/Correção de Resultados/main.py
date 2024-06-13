import pandas as pd
from openpyxl import load_workbook
from os import system

system("cls")

# Caminhos das planilhas
caminho_planilha_origem = 'resultado.xlsx'
caminho_planilha_destino = 'Gabarito Base.xlsx'
caminho_planilha_final = 'Resultado Final.xlsx'  # Caminho para o novo arquivo final

# Definir o número de alunos
num_alunos = 20

# Ler os dados da planilha de origem
df_origem = pd.read_excel(caminho_planilha_origem, sheet_name='Respostas Completas')

# Extrair Gabarito da avaliação
gabarito_portugues = df_origem.iloc[0, 3:29].values  # Colunas D até AC (índices 3 até 28)
gabarito_matematica = df_origem.iloc[0, 29:55].values  # Colunas AD até BC (índices 29 até 54)

# Extrair nomes dos alunos
nomes_alunos = df_origem.iloc[1:1+num_alunos, 2].values  # Coluna C (índice 2), começando da linha 3 (índice 2)

# Extrair respostas de português e matemática
respostas_portugues = df_origem.iloc[1:1+num_alunos, 3:29].values  # Colunas D até AC (índices 3 até 28)
respostas_matematica = df_origem.iloc[1:1+num_alunos, 29:55].values  # Colunas AD até BC (índices 29 até 54)

# Carregar a planilha de destino
workbook = load_workbook(caminho_planilha_destino)
worksheet = workbook['Avaliações']  # Usar o nome correto da aba na planilha de destino

# Colar Gabarito de português
for i, gabarito in enumerate(gabarito_portugues):
    worksheet.cell(row=53, column=6+i, value=gabarito)  # Colunas F até AE

# Colar Gabarito de matemática
for i, gabarito in enumerate(gabarito_matematica):
    worksheet.cell(row=53, column=33+i, value=gabarito)  # Colunas AG até BF

# Colar nomes dos alunos
for i, nome in enumerate(nomes_alunos):
    worksheet.cell(row=13+i, column=3, value=nome)  # Coluna C (índice 3)

# Colar respostas de português
for i, respostas in enumerate(respostas_portugues):
    for j, resposta in enumerate(respostas):
        worksheet.cell(row=13+i, column=6+j, value=resposta)  # Colunas F até AE

# Colar respostas de matemática
for i, respostas in enumerate(respostas_matematica):
    for j, resposta in enumerate(respostas):
        worksheet.cell(row=13+i, column=33+j, value=resposta)  # Colunas AG até BF

# Salvar a planilha final em um novo arquivo
workbook.save(caminho_planilha_final)

print("Dados transferidos com sucesso para o arquivo:", caminho_planilha_final)

