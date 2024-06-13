import pandas as pd
from openpyxl import load_workbook

# Carrega o arquivo Excel
df = pd.read_excel('arquivo.xlsx', header=None, names=['ID', 'Nome'])

# Formata os IDs para sempre ter dois dígitos
df['ID'] = df['ID'].astype(str).str.zfill(2)

# Adiciona o número 3 ao início de cada ID
numero = '3'
df['ID'] = numero + df['ID']

# Separa o nome completo em dois primeiros nomes e restante do nome
df[['Primeiro Nome', 'Segundo Nome', 'Restante Nome']] = df['Nome'].str.split(n=2, expand=True)

# Concatena os dois primeiros nomes em uma coluna
df['NOME'] = df['Primeiro Nome'] + ' ' + df['Segundo Nome']

# Renomeia as colunas e mantém apenas as necessárias
df = df[['ID', 'NOME', 'Restante Nome']]

# Renomeia a coluna do restante do nome
df.rename(columns={'Restante Nome': 'SOBRENOME'}, inplace=True)

# Salva o DataFrame em um arquivo Excel, começando na segunda linha
df.to_excel('resultado_formatado.xlsx', index=False, startrow=1)

# Usar openpyxl para adicionar a linha de título personalizada
wb = load_workbook('resultado_formatado.xlsx')
ws = wb.active
ws.insert_rows(1)  # Insere uma linha no início
ws['A1'] = 'participantes'  # Adiciona o título na célula A1

# Remove a linha vazia entre "participantes" e os cabeçalhos
ws.delete_rows(2)

# Salva novamente o arquivo Excel com a linha de título personalizada
wb.save('resultado_formatado.xlsx')

print("Dados salvos no arquivo 'resultado_formatado.xlsx'.")
