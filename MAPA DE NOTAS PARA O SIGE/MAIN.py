import os
import openpyxl
import pyautogui
import keyboard

# python
# from mouseinfo import mouseInfo
# mouseInfo()
# F6

# os.system('cls')

# PLANILHA = openpyxl.load_workbook('NOTAS.xlsx')
PLANILHA = openpyxl.load_workbook('Exemplo.xlsx')
PAGINA = PLANILHA['NOTAS']

pyautogui.click(82,877, duration=0.25) #Navegador
pyautogui.doubleClick(16, 109, duration=0.25)
pyautogui.press('tab')

# CONFIGURAÇÕES:
serie = 0 # 0 = 6º e 7º ano, 1 = 8º e 9º ano, 2 = EJA I e II, 3 = EJA III e IV
coluna_inicial = 1 # Defalt: 1. 1ª coluna da planilha a ser preenchida
linha_inicial = 5 # Defalt: 5. 1ª linha da planilha a ser preenchida

# Lista das colunas a serem preenchidas
if serie == 0: #6º e 7º ano
    Lista = {10: 'Art.', 28: 'Ciên.', 16: 'E.F.', 34: 'Relig.', 58: 'É.C.', 46: 'Geo.', 40: 'His.', 52: 'Ing.', 4: 'Port.', 22: 'Mat.', 64: 'P.I.', 70: 'Proj.C.'}
elif serie == 1: #8º e 9º anos
    Lista = {10: 'Art.', 28: 'Ciên.', 16: 'E.F.', 34: 'Relig.', 58: 'É.C.', 46: 'Geo.', 40: 'His.', 52: 'Ing.', 4: 'Port.', 22: 'Mat.', 64: 'P.I.'}
elif serie == 2: #EJA I e II
    Lista = {10: 'Art.', 28: 'Ciên.', 16: 'E.F.', 34: 'Relig.', 46: 'Geo.', 40: 'His.', 4: 'Port.', 22: 'Mat.'}
elif serie == 3: #EJA III e IV
    Lista = {10: 'Art.', 28: 'Ciên.', 16: 'E.F.', 34: 'Relig.', 46: 'Geo.', 40: 'His.', 52: 'Ing.', 4: 'Port.', 22: 'Mat.'}

adicionarcoluna = 0 # Defalt: 0. Adiciona numeração ou subtrai da key da Lista anterior:
Lista = {key + adicionarcoluna: value for key, value in Lista.items()}

colunas = list(Lista.keys())

def check_for_esc_key():
    if keyboard.is_pressed('esc'):
        print("\n\nTecla Esc pressionada. Encerrando o script.\n")
        exit()

def converter_materia(colunas):
    dicionario = Lista.copy()
    materia = {k: dicionario[k] for k in colunas}
    return materia

materias = converter_materia(colunas)
x = 0
contador = 1  # Inicia o contador
for linha in PAGINA.iter_rows(min_row=linha_inicial, min_col=coluna_inicial):
    if linha[0].value is not None:
        print(f'|{contador:>2}| {" ".join(linha[0].value.split()[:2]):<20} ', end='|')
        contador += 1  # Incrementa o contador
        for coluna in colunas:
            check_for_esc_key()  # Verifique se a tecla Esc foi pressionada
            valor_celula = str(linha[coluna].value)
            nome_materia = materias[coluna]
            print(f' {nome_materia}: {valor_celula:<4} ', end='|')
            x = 0 if x == len(materias.keys()) -1 else x + 1
            print('') if x == 0 else None
            if valor_celula != "#DIV/0!":
                pyautogui.write(valor_celula.replace('.', ','))
            # Se o valor for "#DIV/0!", deixe em branco
            pyautogui.press('tab')
    else:
        break
